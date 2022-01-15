from openpyxl import Workbook, load_workbook
import logging

LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"

logging.basicConfig(filename = "log.txt", filemode = "w", format = LOG_FORMAT, level = logging.INFO)
logger = logging.getLogger()

def get_file():
    month = input("Would you like to get data on January or March data?")
    try:
        if month.lower() == "january" or month.lower() == "march":
            get_data(month)
        else:
            raise TypeError()
    except:
        print("Please input a correct value: 'January' or 'March'.")
        get_file()

def get_data(month):
    if month.lower() == "january":
        wb = load_workbook("january_2018.xlsx")
        ws = wb["Summary Rolling MoM"]
        
        calls = ws["B12"].value
        abandon = ws["C12"].value
        fcr = ws["D12"].value
        dsat = ws["E12"].value
        csat = ws["F12"].value
        
        logger.info("January Summary Data:")
        logger.info(f"Calls Offered: {calls}")
        logger.info(f"Abandoned after 30 seconds: {abandon * 100}%")
        logger.info(f"FCR: {fcr * 100}%")
        logger.info(f"DSAT: {dsat * 100}%")
        logger.info(f"CSAT: {csat * 100}%")
        logger.info("**************************************")
        
        ws = wb["VOC Rolling MoM"]
        
        size = ws["C3"].value
        promoters = ws["C4"].value
        gb_promo = "GOOD" if promoters > 200 else "BAD"
        passives = ws["C6"].value
        gb_pass = "GOOD" if passives > 100 else "BAD"
        dectractors = ws["C8"].value
        gb_dec = "GOOD" if dectractors > 100 else "BAD"
        nps = ws["C13"].value
        sat = ws["C16"].value
        dsat = ws["C19"].value
        
        logger.info("January VOC Data:")
        logger.info(f"Base Size: {size}")
        logger.info(f"Promoters: {promoters} {gb_promo}")
        logger.info(f"Passives: {passives} {gb_pass}")
        logger.info(f"Dectractors: {dectractors} {gb_dec}")
        logger.info(f"Overal NPS %: {nps * 100}%")
        logger.info(f"Sat with Agent %: {sat * 100}%")
        logger.info(f"DSat with Agent %: {dsat * 100}%")
        
        print("All done! Check log.txt for data.")
    elif month.lower() == "march":
        wb = load_workbook("march_2018.xlsx")
        ws = wb["Summary Rolling MoM"]
        
        calls = ws["B13"].value
        abandon = ws["C13"].value
        fcr = ws["D13"].value
        dsat = ws["E13"].value
        csat = ws["F13"].value
        
        logger.info("March Data:")
        logger.info(f"Calls Offered: {calls}")
        logger.info(f"Abandoned after 30 seconds: {abandon * 100}%")
        logger.info(f"FCR: {fcr * 100}%")
        logger.info(f"DSAT: {dsat * 100}%")
        logger.info(f"CSAT: {csat * 100}%")
        
        ws = wb["VOC Rolling MoM"]
        
        size = ws["B3"].value
        promoters = ws["B4"].value
        gb_promo = "GOOD" if promoters > 200 else "BAD"
        passives = ws["B6"].value
        gb_pass = "GOOD" if passives > 100 else "BAD"
        dectractors = ws["B8"].value
        gb_dec = "GOOD" if dectractors > 100 else "BAD"
        nps = ws["B13"].value
        sat = ws["B16"].value
        dsat = ws["B19"].value
        
        logger.info("March VOC Data:")
        logger.info(f"Base Size: {size}")
        logger.info(f"Promoters: {promoters} {gb_promo}")
        logger.info(f"Passives: {passives} {gb_pass}")
        logger.info(f"Dectractors: {dectractors} {gb_dec}")
        logger.info(f"Overal NPS %: {nps * 100}%")
        logger.info(f"Sat with Agent %: {sat * 100}%")
        logger.info(f"DSat with Agent %: {dsat * 100}%")
        
        print("All done! Check log.txt for data.")
    else:
        print("File not found.")
    
get_file()